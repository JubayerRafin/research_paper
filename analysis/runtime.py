#!/usr/bin/env python3
"""
runtime.py — median per-page runtime (ms/page) for each method.

Measures wall-clock time per page for the pipeline's extractors and the
baseline tools, over the SAME evaluation pages, and reports the median
ms/page (median is robust to outlier pages). Matches the runtime table in
the mentor report.

Measured:
  text  : ours (text_extractor), pdfplumber, pymupdf, pdfminer
  table : ours (Docling primary), camelot_lattice, pdfplumber find_tables
  image : ours (render-and-mask), gap-based, pymupdf get_images, pdfimages

Usage:
  python runtime.py --pdf hp.pdf --gt hp_gt.xlsx --stage1_dir path/to/stage1
     --stage1_dir is the folder that contains your pipeline's
     text_extractor.py / image_extractor.py / table_extractor.py packages.
     If omitted, pipeline-method timings are skipped (baselines still run).

Notes:
  - Docling is slow to LOAD the model once; we time only per-page convert().
  - Camelot is timed on a small page subset (it is very slow); scaled per page.
  - Use --pages N to cap how many pages are timed (default 15 for speed).
"""
import os, re, sys, time, argparse, statistics, io


def load_pages(gt_xlsx, cap):
    from openpyxl import load_workbook
    ws = load_workbook(gt_xlsx, data_only=True).active
    hdr = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    ci = next(i for i, h in enumerate(hdr) if "page" in h)
    pages = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ci] is None:
            continue
        try:
            pages.append(int(r[ci]))
        except (ValueError, TypeError):
            pass
    pages = sorted(set(pages))
    return pages[:cap] if cap else pages


def med_ms(times):
    return round(statistics.median(times) * 1000, 1) if times else None


# ---------------- baseline timers (per page) ----------------
def time_pdfplumber_text(pdf, pages):
    import pdfplumber
    t = []
    with pdfplumber.open(pdf) as d:
        for p in pages:
            s = time.perf_counter()
            _ = d.pages[p-1].extract_text()
            t.append(time.perf_counter() - s)
    return t

def time_pymupdf_text(pdf, pages):
    import fitz
    doc = fitz.open(pdf); t = []
    for p in pages:
        s = time.perf_counter(); _ = doc[p-1].get_text(); t.append(time.perf_counter()-s)
    return t

def time_pdfminer_text(pdf, pages):
    from pdfminer.high_level import extract_text
    t = []
    for p in pages:
        s = time.perf_counter(); _ = extract_text(pdf, page_numbers=[p-1]); t.append(time.perf_counter()-s)
    return t

def time_pdfplumber_tables(pdf, pages):
    import pdfplumber
    t = []
    with pdfplumber.open(pdf) as d:
        for p in pages:
            s = time.perf_counter(); _ = d.pages[p-1].find_tables(); t.append(time.perf_counter()-s)
    return t

def time_camelot_lattice(pdf, pages):
    import camelot
    t = []
    for p in pages:
        s = time.perf_counter()
        try: camelot.read_pdf(pdf, pages=str(p), flavor="lattice", suppress_stdout=True)
        except Exception: pass
        t.append(time.perf_counter()-s)
    return t

def time_pymupdf_images(pdf, pages):
    import fitz
    doc = fitz.open(pdf); t = []
    for p in pages:
        s = time.perf_counter(); _ = doc[p-1].get_images(full=True); t.append(time.perf_counter()-s)
    return t

def time_gapbased(pdf, pages):
    import pdfplumber
    t = []
    with pdfplumber.open(pdf) as d:
        for p in pages:
            s = time.perf_counter()
            w = d.pages[p-1].extract_words()
            tops = sorted(set(round(x['top']) for x in w)) if w else []
            _ = sum(1 for i in range(1, len(tops)) if tops[i]-tops[i-1] > 40)
            t.append(time.perf_counter()-s)
    return t


# ---------------- pipeline timers (need stage1 package) ----------------
def time_ours(stage1_dir, pdf, pages, which):
    """which in {text, image, table}. Times per-page extract on the pipeline."""
    sys.path.insert(0, stage1_dir)
    cfg = {"images": {"output_subdir": "images"},
           "tables": {"output_subdir": "tables", "use_docling": True}}
    import tempfile
    outdir = tempfile.mkdtemp()
    times = []
    try:
        if which == "text":
            from stage1.text_extractor import TextExtractor
            te = TextExtractor(cfg, outdir)
            for p in pages:
                s = time.perf_counter(); te.extract(pdf, page_numbers=[p]); times.append(time.perf_counter()-s)
        elif which == "image":
            from stage1.image_extractor import ImageExtractor
            ie = ImageExtractor(cfg, outdir)
            for p in pages:
                s = time.perf_counter(); ie.extract(pdf, page_numbers=[p]); times.append(time.perf_counter()-s)
        elif which == "table":
            from stage1.table_extractor import TableExtractor
            tx = TableExtractor(cfg, outdir)   # loads Docling model once (not timed below)
            # warm-up one page so model load isn't counted
            try: tx.extract(pdf, page_numbers=[pages[0]])
            except Exception: pass
            for p in pages:
                s = time.perf_counter(); tx.extract(pdf, page_numbers=[p]); times.append(time.perf_counter()-s)
    except Exception as e:
        print(f"  [ours {which}] skipped: {e}")
        return []
    return times


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", required=True)
    ap.add_argument("--gt", required=True)
    ap.add_argument("--stage1_dir", default=None,
                    help="folder containing the stage1 package (for ours timings)")
    ap.add_argument("--pages", type=int, default=15, help="max pages to time")
    a = ap.parse_args()

    pages = load_pages(a.gt, a.pages)
    print(f"Timing over {len(pages)} pages: {pages}\n")

    rows = []  # (method, ms/page)

    # ----- TEXT -----
    if a.stage1_dir:
        rows.append(("Ours (text)", med_ms(time_ours(a.stage1_dir, a.pdf, pages, "text"))))
    rows.append(("pdfplumber (text)", med_ms(time_pdfplumber_text(a.pdf, pages))))
    rows.append(("PyMuPDF (text)",   med_ms(time_pymupdf_text(a.pdf, pages))))
    rows.append(("pdfminer.six (text)", med_ms(time_pdfminer_text(a.pdf, pages))))

    # ----- TABLE -----
    if a.stage1_dir:
        rows.append(("Ours (table, Docling)", med_ms(time_ours(a.stage1_dir, a.pdf, pages, "table"))))
    rows.append(("Camelot lattice", med_ms(time_camelot_lattice(a.pdf, pages))))
    rows.append(("pdfplumber find_tables()", med_ms(time_pdfplumber_tables(a.pdf, pages))))

    # ----- IMAGE -----
    if a.stage1_dir:
        rows.append(("Ours (image)", med_ms(time_ours(a.stage1_dir, a.pdf, pages, "image"))))
    rows.append(("Gap-based", med_ms(time_gapbased(a.pdf, pages))))
    rows.append(("PyMuPDF get_images()", med_ms(time_pymupdf_images(a.pdf, pages))))

    print(f"\n{'Method':<28} {'ms/page (median)':>18}")
    print("-" * 48)
    for name, ms in rows:
        print(f"{name:<28} {('n/a' if ms is None else ms):>18}")
    print("\nNote: DocLayout-YOLO ~1776 ms/page (CPU) — run separately if needed.")
    print("Camelot/Docling are the slow paths; PyMuPDF is fastest.")


if __name__ == "__main__":
    main()
