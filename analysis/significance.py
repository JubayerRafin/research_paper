#!/usr/bin/env python3
"""
significance.py — paired bootstrap significance testing.

For each metric (image / table / text), compares OURS against each baseline
using a paired bootstrap over PAGES: resample the evaluated pages with
replacement B times, recompute the metric difference (ours - baseline) each
time, and derive a two-sided p-value and 95% CI.

Per-page metric used:
  - image / table : per-page F1 of detection counts (count-based)
  - text          : per-page token-F1

A positive mean difference means OURS is better. p < 0.05 => significant.

Inputs are the SAME artifacts used by the scorers:
  --gt hp_gt.xlsx          (image/table counts)
  --output output          (ours: images/ and tables/)
  --pdf hp.pdf             (to run baselines)
  --gt_dir text_gt_50      (text GT, for text significance)
  --md output/xxx.md       (ours text)

Usage:
  python significance.py --metric image --gt hp_gt.xlsx --output output --pdf hp.pdf
  python significance.py --metric table --gt hp_gt.xlsx --output output --pdf hp.pdf
  python significance.py --metric text  --gt_dir text_gt_50 --md output/xxx.md --pdf hp.pdf
"""
import os, re, glob, argparse, random
from collections import defaultdict, Counter

B = 10000          # bootstrap iterations
random.seed(42)


# ---------- shared ----------
def f1_from_counts(pred, gt):
    tp = min(pred, gt); fp = max(0, pred - gt); fn = max(0, gt - pred)
    P = tp / (tp + fp) if (tp + fp) else (1.0 if gt == 0 else 0.0)
    R = tp / (tp + fn) if (tp + fn) else (1.0 if pred == 0 else 0.0)
    if gt == 0 and pred == 0:
        return 1.0
    return 2 * P * R / (P + R) if (P + R) else 0.0


def token_f1(ref, hyp):
    r = Counter(ref.split()); h = Counter(hyp.split())
    tp = sum((r & h).values())
    P = tp / sum(h.values()) if sum(h.values()) else 0.0
    R = tp / sum(r.values()) if sum(r.values()) else 0.0
    return 2 * P * R / (P + R) if (P + R) else 0.0


def paired_bootstrap(ours_pp, base_pp):
    """ours_pp, base_pp: lists of per-page scores (aligned by page)."""
    n = len(ours_pp)
    diffs = [o - b for o, b in zip(ours_pp, base_pp)]
    obs = sum(diffs) / n
    boot = []
    idx = list(range(n))
    for _ in range(B):
        sample = [diffs[random.choice(idx)] for _ in range(n)]
        boot.append(sum(sample) / n)
    boot.sort()
    lo = boot[int(0.025 * B)]
    hi = boot[int(0.975 * B)]
    # two-sided p: fraction of bootstrap means on the opposite side of 0
    if obs >= 0:
        p = 2 * sum(1 for b in boot if b <= 0) / B
    else:
        p = 2 * sum(1 for b in boot if b >= 0) / B
    p = min(p, 1.0)
    return obs, lo, hi, p


# ---------- GT / counts ----------
def load_gt_counts(xlsx, key):
    from openpyxl import load_workbook
    ws = load_workbook(xlsx, data_only=True).active
    hdr = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    ci_page = next(i for i, h in enumerate(hdr) if "page" in h)
    ci = next(i for i, h in enumerate(hdr) if key in h)
    gt = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ci_page] is None:
            continue
        try:
            gt[int(r[ci_page])] = int(r[ci] or 0)
        except (ValueError, TypeError):
            pass
    return gt


def count_files(folder, pattern):
    c = defaultdict(int)
    for f in glob.glob(os.path.join(folder, pattern)):
        m = re.search(r"_p(\d+)_", os.path.basename(f))
        if m:
            c[int(m.group(1))] += 1
    return c


# ---------- baseline count providers ----------
def ours_image_counts(output):
    return count_files(os.path.join(output, "images"), "image_p*.png")

def ours_table_counts(output):
    return count_files(os.path.join(output, "tables"), "table_p*.json")

def base_image_counts(name, pdf, pages):
    if name == "pymupdf":
        import fitz; doc = fitz.open(pdf)
        return {p: len(doc[p-1].get_images(full=True)) for p in pages}
    if name == "gap-based":
        import pdfplumber; c = {}
        with pdfplumber.open(pdf) as d:
            for p in pages:
                w = d.pages[p-1].extract_words()
                tops = sorted(set(round(x['top']) for x in w)) if w else []
                c[p] = sum(1 for i in range(1, len(tops)) if tops[i]-tops[i-1] > 40)
        return c
    raise ValueError(name)

def base_table_counts(name, pdf, pages):
    if name == "pdfplumber":
        import pdfplumber; c = {}
        with pdfplumber.open(pdf) as d:
            for p in pages:
                try: c[p] = len(d.pages[p-1].find_tables())
                except: c[p] = 0
        return c
    if name in ("camelot_lattice", "camelot_stream"):
        import camelot
        flavor = name.split("_")[1]
        c = defaultdict(int)
        try:
            for t in camelot.read_pdf(pdf, pages=",".join(map(str, pages)),
                                      flavor=flavor, suppress_stdout=True):
                c[int(t.page)] += 1
        except Exception as e:
            print(f"  camelot {flavor} error: {e}")
        return {p: c.get(p, 0) for p in pages}
    raise ValueError(name)


# ---------- text ----------
def load_text_gt(gt_dir):
    gt = {}
    for f in glob.glob(os.path.join(gt_dir, "p*.txt")):
        m = re.search(r"p(\d+)\.txt", os.path.basename(f))
        if m:
            gt[int(m.group(1))] = re.sub(r"\s+", " ",
                                         open(f, encoding="utf-8").read()).strip()
    return gt

def ours_text(md):
    raw = open(md, encoding="utf-8").read()
    parts = re.split(r"<!--\s*page:\s*(\d+)\s*-->", raw)
    pages = {}
    for i in range(1, len(parts), 2):
        body = parts[i+1] if i+1 < len(parts) else ""
        body = re.sub(r"```.*?```", " ", body, flags=re.DOTALL)
        body = re.sub(r"!\[.*?\]\(.*?\)", " ", body)
        body = re.sub(r"^#{1,6}\s*", "", body, flags=re.MULTILINE)
        pages[int(parts[i])] = re.sub(r"\s+", " ", body).strip()
    return pages

def base_text(name, pdf, pages):
    out = {}
    if name == "pdfplumber":
        import pdfplumber
        with pdfplumber.open(pdf) as d:
            for p in pages:
                try: out[p] = re.sub(r"\s+"," ",d.pages[p-1].extract_text() or "").strip()
                except: out[p] = ""
    elif name == "pymupdf":
        import fitz; doc = fitz.open(pdf)
        for p in pages:
            try: out[p] = re.sub(r"\s+"," ",doc[p-1].get_text()).strip()
            except: out[p] = ""
    elif name == "pdfminer":
        from pdfminer.high_level import extract_text
        for p in pages:
            try: out[p] = re.sub(r"\s+"," ",extract_text(pdf,page_numbers=[p-1])).strip()
            except: out[p] = ""
    return out



def load_doclayout_counts(path, idx):
    """Load cached DocLayout per-page counts from JSON produced by
    baseline_doclayout.py --dump. idx=0 -> figures, idx=1 -> tables."""
    import json
    d = json.load(open(path, encoding="utf-8"))
    return {int(k): v[idx] for k, v in d.items()}


def report(metric, comparisons):
    print(f"\n=== SIGNIFICANCE — {metric.upper()} (paired bootstrap, B={B}) ===")
    print(f"{'comparison':<28} {'mean diff':>10} {'95% CI':>22} {'p':>8}  result")
    for label, (obs, lo, hi, p) in comparisons:
        sig = "significant" if p < 0.05 else "not sig."
        print(f"{label:<28} {obs:>+10.4f}  [{lo:>+7.3f}, {hi:>+7.3f}]  {p:>7.4f}  {sig}")
    print()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--metric", required=True, choices=["image", "table", "text"])
    ap.add_argument("--gt"); ap.add_argument("--output"); ap.add_argument("--pdf")
    ap.add_argument("--gt_dir"); ap.add_argument("--md")
    ap.add_argument("--doclayout", help="cached doclayout counts .json (optional)")
    a = ap.parse_args()

    if a.metric in ("image", "table"):
        key = "image" if a.metric == "image" else "table"
        gt = load_gt_counts(a.gt, key)
        pages = sorted(gt)
        ours_counts = (ours_image_counts if a.metric == "image"
                       else ours_table_counts)(a.output)
        ours_pp = [f1_from_counts(ours_counts.get(p, 0), gt[p]) for p in pages]

        baselines = (["pymupdf", "gap-based"] if a.metric == "image"
                     else ["camelot_lattice", "camelot_stream", "pdfplumber"])
        comps = []
        for name in baselines:
            bc = (base_image_counts if a.metric == "image"
                  else base_table_counts)(name, a.pdf, pages)
            base_pp = [f1_from_counts(bc.get(p, 0), gt[p]) for p in pages]
            comps.append((f"Ours vs {name}", paired_bootstrap(ours_pp, base_pp)))
        if a.doclayout:
            didx = 0 if a.metric == "image" else 1
            dc = load_doclayout_counts(a.doclayout, didx)
            d_pp = [f1_from_counts(dc.get(p, 0), gt[p]) for p in pages]
            comps.append(("Ours vs DocLayout-YOLO",
                          paired_bootstrap(ours_pp, d_pp)))
        report(a.metric, comps)

    else:  # text
        gt = load_text_gt(a.gt_dir)
        pages = sorted(p for p in gt if gt[p].strip())
        omap = ours_text(a.md)
        ours_pp = [token_f1(gt[p], omap.get(p, "")) for p in pages]
        comps = []
        for name in ["pdfplumber", "pymupdf", "pdfminer"]:
            bmap = base_text(name, a.pdf, pages)
            base_pp = [token_f1(gt[p], bmap.get(p, "")) for p in pages]
            comps.append((f"Ours vs {name}", paired_bootstrap(ours_pp, base_pp)))
        report("text", comps)


if __name__ == "__main__":
    main()
