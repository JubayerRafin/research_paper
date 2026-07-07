#!/usr/bin/env python3
"""
baseline_doclayout.py — DocLayout-YOLO as a SOTA baseline for IMAGE and TABLE.

Renders each benchmark page at 150 DPI, runs DocLayout-YOLO, counts 'figure'
and 'table' detections per page, and scores P/R/F1 against hp_gt.xlsx — the
same count-based metric and the same pages as the pipeline scorer.

SETUP:
  pip install doclayout-yolo pymupdf openpyxl pillow
  # weights auto-download on first run (needs internet once), then offline.

RUN (from the HP_evaluation folder):
  python baseline\\baseline_doclayout.py

NOTES:
  - 'figure' detections -> image metric; 'table' detections -> table metric.
  - DocLayout table detection locates regions only; it extracts NO cell
    content. Report its table score as DETECTION-ONLY.
  - Prints per-page counts so you can sanity-check before trusting the score.
"""
import os, re, sys, io, argparse
from openpyxl import load_workbook

# ---- defaults (override with CLI args) ----
DEFAULT_PDF = "hp-e877-series-user-guide.pdf"
DEFAULT_GT  = "ground_truth/hp_gt.xlsx"
DPI  = 150
CONF = 0.25
WEIGHTS = "doclayout_yolo_docstructbench_imgsz1024.pt"


def load_gt(xlsx):
    """Read GT by COLUMN NAME (robust to column order). Expects headers
    containing 'page', 'table', 'image'."""
    ws = load_workbook(xlsx, data_only=True).active
    hdr = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    ci_page = next(i for i, h in enumerate(hdr) if "page" in h)
    ci_tbl  = next(i for i, h in enumerate(hdr) if "table" in h)
    ci_img  = next(i for i, h in enumerate(hdr) if "image" in h)
    gt = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ci_page] is None:
            continue
        try:
            p = int(row[ci_page])
        except (ValueError, TypeError):
            continue
        gt[p] = dict(tables=int(row[ci_tbl] or 0),
                     images=int(row[ci_img] or 0))
    return gt


def prf(tp, fp, fn):
    P = tp / max(tp + fp, 1)
    R = tp / max(tp + fn, 1)
    F = 2 * P * R / max(P + R, 1e-9)
    return round(P, 4), round(R, 4), round(F, 4)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", default=DEFAULT_PDF)
    ap.add_argument("--gt",  default=DEFAULT_GT)
    ap.add_argument("--conf", type=float, default=CONF)
    ap.add_argument("--dump", help="save per-page counts to this JSON (for significance)")
    args = ap.parse_args()

    gt = load_gt(args.gt)
    pages = sorted(gt)

    # ---- model ----
    try:
        from doclayout_yolo import YOLOv10
    except ImportError:
        print("ERROR: pip install doclayout-yolo pymupdf openpyxl pillow")
        sys.exit(1)

    import fitz
    from PIL import Image

    weights = WEIGHTS
    if not os.path.exists(weights):
        try:
            from huggingface_hub import hf_hub_download
            weights = hf_hub_download(
                repo_id="juliozhao/DocLayout-YOLO-DocStructBench",
                filename="doclayout_yolo_docstructbench_imgsz1024.pt")
            print(f"[weights] downloaded to {weights}")
        except Exception as e:
            print(f"ERROR getting weights: {e}")
            sys.exit(1)

    model = YOLOv10(weights)
    print(f"[model] loaded. classes: {model.names}")

    FIGURE_LABELS = {"figure"}
    TABLE_LABELS  = {"table"}

    doc = fitz.open(args.pdf)
    pred = {}
    print("\n[per-page detections]  page: figures / tables   (GT: img / tbl)")
    for p in pages:
        page = doc[p - 1]
        pix = page.get_pixmap(dpi=DPI)
        img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
        res = model.predict(img, imgsz=1024, conf=args.conf, verbose=False)[0]
        figs = tbls = 0
        for cls_id in res.boxes.cls.tolist():
            name = model.names[int(cls_id)].lower()
            if name in FIGURE_LABELS:
                figs += 1
            elif name in TABLE_LABELS:
                tbls += 1
        pred[p] = (figs, tbls)
        print(f"  p{p:>3}: {figs} / {tbls}   "
              f"(GT: {gt[p]['images']} / {gt[p]['tables']})")

    if args.dump:
        import json
        json.dump({str(p): list(pred[p]) for p in pages},
                  open(args.dump, "w", encoding="utf-8"))
        print(f"[dump] per-page counts written to {args.dump}")

    def score(idx, gt_key, label):
        tp = fp = fn = 0
        for p in pages:
            pv = pred[p][idx]; gv = gt[p][gt_key]
            tp += min(pv, gv); fp += max(pv - gv, 0); fn += max(gv - pv, 0)
        P, R, F = prf(tp, fp, fn)
        print(f"  {label:<34} P={P:.3f} R={R:.3f} F1={F:.3f}  "
              f"(tp={tp} fp={fp} fn={fn})")

    print("\n" + "=" * 62)
    print("DocLayout-YOLO (SOTA reference) — same GT/pages as pipeline")
    print("=" * 62)
    score(0, "images", "DocLayout-YOLO (figures)")
    score(1, "tables", "DocLayout-YOLO (tables, DETECTION-ONLY)")

    print("\nCurrent pipeline results for reference:")
    print("  IMAGE render-and-mask   P=1.000 R=0.750 F1=0.857")
    print("  TABLE Docling+fallback  P=0.875 R=0.972 F1=0.921  "
          "(+ cell-content F1=0.961)")
    print("\nNote: DocLayout table detection locates regions only and extracts")
    print("no cell content. It runs on CPU but at ~1776 ms/page (~9x slower")
    print("than our 200 ms/page image extractor) and requires the DL runtime")
    print("plus a one-time model-weights download.")


if __name__ == "__main__":
    main()