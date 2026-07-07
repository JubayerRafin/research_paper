#!/usr/bin/env python3
"""
baseline_table.py — TABLE detection baselines, count-based P/R/F1.

Baselines:
  - camelot_lattice : Camelot flavor='lattice'
  - camelot_stream  : Camelot flavor='stream'
  - pdfplumber      : page.find_tables()

Scores each against hp_gt.xlsx (gt_tables), same pages, same count metric.

Usage:
  python baseline_table.py --pdf hp.pdf --gt hp_gt.xlsx
"""
import argparse
from collections import defaultdict


def load_gt(xlsx, key="table"):
    from openpyxl import load_workbook
    ws = load_workbook(xlsx, data_only=True).active
    hdr = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    ci_page = next(i for i,h in enumerate(hdr) if "page" in h)
    ci = next(i for i,h in enumerate(hdr) if key in h)
    gt = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ci_page] is None: continue
        try: p=int(r[ci_page])
        except: continue
        gt[p]=int(r[ci] or 0)
    return gt


def prf(counts, gt):
    TP=FP=FN=0
    for p in gt:
        g=gt[p]; e=counts.get(p,0)
        TP+=min(e,g); FP+=max(0,e-g); FN+=max(0,g-e)
    P=TP/(TP+FP) if TP+FP else 0.0
    R=TP/(TP+FN) if TP+FN else 0.0
    F=2*P*R/(P+R) if P+R else 0.0
    return P,R,F,TP,FP,FN


def count_camelot(pdf, pages, flavor):
    import camelot
    c=defaultdict(int)
    page_str=",".join(str(p) for p in pages)
    try:
        tables=camelot.read_pdf(pdf,pages=page_str,flavor=flavor,suppress_stdout=True)
        for t in tables:
            c[int(t.page)]+=1
    except Exception as e:
        print(f"  [camelot {flavor}] error: {e}")
    return {p:c.get(p,0) for p in pages}


def count_pdfplumber(pdf, pages):
    import pdfplumber
    c={}
    with pdfplumber.open(pdf) as doc:
        for p in pages:
            try: c[p]=len(doc.pages[p-1].find_tables())
            except: c[p]=0
    return c


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--pdf",required=True); ap.add_argument("--gt",required=True)
    a=ap.parse_args()
    gt=load_gt(a.gt,"table"); pages=sorted(gt)
    print(f"TABLE baselines | {len(pages)} pages\n")
    for name,fn in [("camelot_lattice",lambda p,pg:count_camelot(p,pg,"lattice")),
                    ("camelot_stream", lambda p,pg:count_camelot(p,pg,"stream")),
                    ("pdfplumber",     count_pdfplumber)]:
        counts=fn(a.pdf,pages)
        P,R,F,TP,FP,FN=prf(counts,gt)
        print(f"  {name:16} P={P:.3f} R={R:.3f} F1={F:.3f}  (TP={TP} FP={FP} FN={FN})")
    print()

if __name__=="__main__": main()
