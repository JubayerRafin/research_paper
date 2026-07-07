#!/usr/bin/env python3
"""
baseline_image.py — IMAGE extraction baselines, count-based P/R/F1.

Baselines:
  - pymupdf   : fitz page.get_images() (embedded raster objects)
  - pdfimages : Poppler pdfimages -list (embedded images)
  - gapbased  : simple text-gap detection (earlier method approximation)

Scores each against hp_gt.xlsx (gt_images), same 50 pages, same count metric
as the pipeline scorer. Reports P/R/F1 per method.

Usage:
  python baseline_image.py --pdf hp.pdf --gt hp_gt.xlsx
"""
import os, re, argparse, subprocess
from collections import defaultdict


def load_gt(xlsx, key="image"):
    from openpyxl import load_workbook
    ws = load_workbook(xlsx, data_only=True).active
    hdr = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    ci_page = next(i for i,h in enumerate(hdr) if "page" in h)
    ci = next(i for i,h in enumerate(hdr) if key in h)
    gt = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ci_page] is None: continue
        try: p = int(r[ci_page])
        except: continue
        gt[p] = int(r[ci] or 0)
    return gt


def prf(counts, gt):
    TP=FP=FN=0
    for p in gt:
        g=gt[p]; e=counts.get(p,0)
        TP+=min(e,g); FP+=max(0,e-g); FN+=max(0,g-e)
    P=TP/(TP+FP) if TP+FP else 0.0
    R=TP/(TP+FN) if TP+FN else 0.0
    F=2*P*R/(P+R) if P+R else 0.0
    return P,R,F,TP,FP,FN


def count_pymupdf(pdf, pages):
    import fitz
    doc=fitz.open(pdf); c={}
    for p in pages:
        try: c[p]=len(doc[p-1].get_images(full=True))
        except: c[p]=0
    return c


def count_pdfimages(pdf, pages):
    """Use poppler pdfimages -list; count images per page. Falls back to 0s
    if pdfimages binary not installed."""
    c=defaultdict(int)
    try:
        out=subprocess.run(["pdfimages","-list",pdf],capture_output=True,text=True).stdout
    except FileNotFoundError:
        print("  [pdfimages] poppler not installed — skipping")
        return {p:0 for p in pages}
    for line in out.splitlines():
        m=re.match(r"\s*(\d+)\s+\d+", line)
        if m:
            pg=int(m.group(1))
            c[pg]+=1
    return {p:c.get(p,0) for p in pages}


def count_gapbased(pdf, pages):
    """Approximate earlier gap-based method: detect vertical gaps between text
    lines wider than a threshold as image regions."""
    import pdfplumber
    c={}
    with pdfplumber.open(pdf) as doc:
        for p in pages:
            try:
                page=doc.pages[p-1]
                words=page.extract_words()
                if not words: c[p]=0; continue
                tops=sorted(set(round(w['top']) for w in words))
                gaps=0; GAP_MIN=40
                for i in range(1,len(tops)):
                    if tops[i]-tops[i-1]>GAP_MIN: gaps+=1
                c[p]=gaps
            except: c[p]=0
    return c


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--pdf",required=True); ap.add_argument("--gt",required=True)
    a=ap.parse_args()
    gt=load_gt(a.gt,"image"); pages=sorted(gt)
    print(f"IMAGE baselines | {len(pages)} pages\n")
    for name,fn in [("pymupdf",count_pymupdf),("pdfimages",count_pdfimages),
                    ("gap-based",count_gapbased)]:
        counts=fn(a.pdf,pages)
        P,R,F,TP,FP,FN=prf(counts,gt)
        print(f"  {name:12} P={P:.3f} R={R:.3f} F1={F:.3f}  (TP={TP} FP={FP} FN={FN})")
    print()

if __name__=="__main__": main()
